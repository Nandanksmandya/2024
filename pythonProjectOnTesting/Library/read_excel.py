import openpyxl

from Library.configure import Configuration


class ReadExcel:
    @staticmethod
    def chage_theme_locators():
        path = Configuration.LOCATORS_PATH
        wb = openpyxl.load_workbook(path)
        ws = wb["locators_details"]
        rows = ws.rows
        header = next(rows)
        data = {}
        for row in ws.rows:
            # key = row[0].value  # Use the first column as the key
            # value = row[1:]  # Use the remaining columns as the value
            # data[key] = value
            data[row[0].value] = (row[1].value, row[2].value)
        return data

    @staticmethod
    def login_details():
        # path = '../Test_data/fb_data.xlsx'  # Replace with the correct path
        path = Configuration.LOCATORS_PATH
        wb = openpyxl.load_workbook(path)
        ws = wb["login_"]
        l_ = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            l_.append((row[0], row[1], row[2], row[3], row[4]))
        return l_


r = ReadExcel()
data = r.login_details()

print(data)
